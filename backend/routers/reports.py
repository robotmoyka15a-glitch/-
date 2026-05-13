"""
WashControl — роутер отчётов
Дневные / недельные / месячные отчёты + экспорт Excel
"""

from datetime import date, datetime, timedelta
from typing import Optional
from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from backend.database import get_connection
from backend.config import WASH_MODES, PAYMENT_METHODS
from backend.routers.auth import get_current_user, require_admin

router = APIRouter(prefix="/reports", tags=["reports"])


# ── Внутренние функции сбора данных ──────────────────────────────────────────

def _day_report(target_date: str) -> dict:
    conn = get_connection()

    shifts = conn.execute(
        """SELECT s.*, u.full_name, u.username
           FROM shifts s JOIN users u ON s.user_id = u.id
           WHERE s.date = ? ORDER BY s.started_at""",
        (target_date,),
    ).fetchall()

    cars = conn.execute(
        """SELECT c.*, u.full_name
           FROM cars c
           JOIN shifts s ON c.shift_id = s.id
           JOIN users u ON c.user_id = u.id
           WHERE s.date = ? ORDER BY c.arrived_at""",
        (target_date,),
    ).fetchall()

    summary = conn.execute(
        """SELECT
               COUNT(*) as total_cars,
               COALESCE(SUM(amount), 0) as main_revenue,
               COALESCE(SUM(extra_amount), 0) as extra_revenue,
               COALESCE(SUM(amount + extra_amount), 0) as total_revenue,
               SUM(CASE WHEN windows_wiped=1 THEN 1 ELSE 0 END) as wiped_count,
               SUM(CASE WHEN extra_service=1 THEN 1 ELSE 0 END) as extra_count
           FROM cars c
           JOIN shifts s ON c.shift_id = s.id
           WHERE s.date = ?""",
        (target_date,),
    ).fetchone()

    by_mode = conn.execute(
        """SELECT wash_mode, COUNT(*) as cnt,
               COALESCE(SUM(amount), 0) as revenue
           FROM cars c
           JOIN shifts s ON c.shift_id = s.id
           WHERE s.date = ? GROUP BY wash_mode ORDER BY wash_mode""",
        (target_date,),
    ).fetchall()

    by_payment = conn.execute(
        """SELECT payment_method, COUNT(*) as cnt,
               COALESCE(SUM(amount), 0) as revenue
           FROM cars c
           JOIN shifts s ON c.shift_id = s.id
           WHERE s.date = ? GROUP BY payment_method""",
        (target_date,),
    ).fetchall()

    late_shifts = [s for s in shifts if s["is_late"]]

    conn.close()
    return {
        "date":       target_date,
        "shifts":     [dict(s) for s in shifts],
        "cars":       [dict(c) for c in cars],
        "summary":    dict(summary),
        "by_mode":    [dict(r) for r in by_mode],
        "by_payment": [dict(r) for r in by_payment],
        "late_shifts": [dict(s) for s in late_shifts],
        "mode_names": WASH_MODES,
        "payment_names": PAYMENT_METHODS,
    }


def _range_report(date_from: str, date_to: str) -> dict:
    conn = get_connection()
    summary = conn.execute(
        """SELECT
               COUNT(*) as total_cars,
               COALESCE(SUM(amount + extra_amount), 0) as total_revenue,
               COALESCE(AVG(amount + extra_amount), 0) as avg_per_car
           FROM cars c
           JOIN shifts s ON c.shift_id = s.id
           WHERE s.date BETWEEN ? AND ?""",
        (date_from, date_to),
    ).fetchone()

    daily = conn.execute(
        """SELECT s.date,
               COUNT(c.id) as cars,
               COALESCE(SUM(c.amount + c.extra_amount), 0) as revenue
           FROM shifts s
           LEFT JOIN cars c ON c.shift_id = s.id
           WHERE s.date BETWEEN ? AND ?
           GROUP BY s.date ORDER BY s.date""",
        (date_from, date_to),
    ).fetchall()

    conn.close()
    return {
        "date_from": date_from,
        "date_to":   date_to,
        "summary":   dict(summary),
        "daily":     [dict(r) for r in daily],
    }


# ── Excel генерация ───────────────────────────────────────────────────────────

HEADER_FILL   = PatternFill("solid", fgColor="1E3A5F")
HEADER_FONT   = Font(color="FFFFFF", bold=True, size=10)
SUBHEAD_FILL  = PatternFill("solid", fgColor="2D6A9F")
SUBHEAD_FONT  = Font(color="FFFFFF", bold=True, size=9)
EVEN_FILL     = PatternFill("solid", fgColor="EEF4FB")
BORDER        = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left",   vertical="center")


def _set_header(ws, row, cols):
    for c, val in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def _build_excel(report: dict) -> BytesIO:
    wb = openpyxl.Workbook()
    ws_main = wb.active
    ws_main.title = "Журнал машин"

    # ── Лист 1: журнал машин ─────────────────────────────────────────────────
    title_cell = ws_main.cell(row=1, column=1,
        value=f"WashControl — Отчёт за {report['date']}")
    title_cell.font = Font(bold=True, size=14, color="1E3A5F")
    ws_main.merge_cells("A1:L1")
    title_cell.alignment = CENTER

    headers = [
        "#", "Оператор", "Время приезда",
        "Режим мойки", "Оплата", "Сумма (₽)",
        "Доп. услуга", "Доп. тип", "Доп. оплата", "Доп. сумма (₽)",
        "Стёкла", "Примечание",
    ]
    _set_header(ws_main, 2, headers)

    mode_names = report.get("mode_names", WASH_MODES)
    pay_names  = report.get("payment_names", PAYMENT_METHODS)

    for i, car in enumerate(report["cars"], 1):
        row_idx = i + 2
        row_data = [
            i,
            car.get("full_name", ""),
            car.get("arrived_at", "")[:16],
            mode_names.get(car["wash_mode"], str(car["wash_mode"])),
            pay_names.get(car["payment_method"], car["payment_method"]),
            car.get("amount", 0),
            car.get("extra_service_name") or ("Да" if car.get("extra_service") else "Нет"),
            pay_names.get(car.get("extra_payment", ""), "") if car.get("extra_service") else "",
            car.get("extra_amount", 0) if car.get("extra_service") else "",
            car.get("extra_amount", 0),
            "✓" if car.get("windows_wiped") else "—",
            car.get("note") or "",
        ]
        fill = EVEN_FILL if i % 2 == 0 else None
        for c, val in enumerate(row_data, 1):
            cell = ws_main.cell(row=row_idx, column=c, value=val)
            cell.border = BORDER
            cell.alignment = CENTER if c in (1, 11) else LEFT
            if fill:
                cell.fill = fill

    # Ширина колонок
    col_widths = [5, 18, 18, 14, 12, 12, 18, 12, 12, 14, 8, 20]
    for c, w in enumerate(col_widths, 1):
        ws_main.column_dimensions[get_column_letter(c)].width = w

    # ── Лист 2: сводка ────────────────────────────────────────────────────────
    ws_sum = wb.create_sheet("Сводка")
    s = report["summary"]

    summary_rows = [
        ("Всего машин",       s.get("total_cars", 0)),
        ("Выручка (осн.)",    f"{s.get('main_revenue', 0):.2f} ₽"),
        ("Выручка (доп.)",    f"{s.get('extra_revenue', 0):.2f} ₽"),
        ("Итого выручка",     f"{s.get('total_revenue', 0):.2f} ₽"),
        ("С доп. услугой",    s.get("extra_count", 0)),
        ("Протёрты стёкла",   s.get("wiped_count", 0)),
    ]
    ws_sum.column_dimensions["A"].width = 24
    ws_sum.column_dimensions["B"].width = 20

    ws_sum.cell(row=1, column=1, value=f"Сводка за {report['date']}").font = \
        Font(bold=True, size=13, color="1E3A5F")

    for r, (label, val) in enumerate(summary_rows, 3):
        ws_sum.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws_sum.cell(row=r, column=2, value=val)

    # По режимам
    ws_sum.cell(row=11, column=1, value="По режимам").font = Font(bold=True, size=11)
    _set_header(ws_sum, 12, ["Режим", "Машин", "Выручка"])
    for i, row in enumerate(report.get("by_mode", []), 13):
        ws_sum.cell(row=i, column=1,
            value=WASH_MODES.get(row["wash_mode"], str(row["wash_mode"])))
        ws_sum.cell(row=i, column=2, value=row["cnt"])
        ws_sum.cell(row=i, column=3, value=f"{row['revenue']:.2f} ₽")

    # По оплате
    offset = 13 + len(report.get("by_mode", [])) + 2
    ws_sum.cell(row=offset, column=1, value="По типу оплаты").font = Font(bold=True, size=11)
    _set_header(ws_sum, offset + 1, ["Тип оплаты", "Машин", "Выручка"])
    for i, row in enumerate(report.get("by_payment", []), offset + 2):
        ws_sum.cell(row=i, column=1,
            value=PAYMENT_METHODS.get(row["payment_method"], row["payment_method"]))
        ws_sum.cell(row=i, column=2, value=row["cnt"])
        ws_sum.cell(row=i, column=3, value=f"{row['revenue']:.2f} ₽")

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Эндпоинты ─────────────────────────────────────────────────────────────────

@router.get("/day")
def day_report(
    target_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
):
    """JSON-отчёт за день."""
    d = target_date or date.today().isoformat()
    return _day_report(d)


@router.get("/day/excel")
def day_report_excel(
    target_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
):
    """Скачать Excel-отчёт за день."""
    d = target_date or date.today().isoformat()
    report = _day_report(d)
    buf = _build_excel(report)
    filename = f"washcontrol_{d}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/week")
def week_report(current_user: dict = Depends(get_current_user)):
    """Отчёт за текущую неделю."""
    today = date.today()
    monday = today - timedelta(days=today.weekday())
    return _range_report(monday.isoformat(), today.isoformat())


@router.get("/month")
def month_report(
    year: Optional[int] = None,
    month: Optional[int] = None,
    current_user: dict = Depends(get_current_user),
):
    """Отчёт за месяц."""
    today = date.today()
    y = year or today.year
    m = month or today.month
    first = date(y, m, 1).isoformat()
    # последний день месяца
    if m == 12:
        last = date(y + 1, 1, 1) - timedelta(days=1)
    else:
        last = date(y, m + 1, 1) - timedelta(days=1)
    return _range_report(first, last.isoformat())


@router.get("/range")
def range_report(
    date_from: str,
    date_to: str,
    current_user: dict = Depends(get_current_user),
):
    """Отчёт за произвольный период."""
    return _range_report(date_from, date_to)


@router.get("/range/excel")
def range_report_excel(
    date_from: str,
    date_to: str,
    current_user: dict = Depends(get_current_user),
):
    """Excel за произвольный период (склеивает дни)."""
    from io import BytesIO
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Период"
    data = _range_report(date_from, date_to)

    ws.cell(row=1, column=1,
        value=f"WashControl {date_from} — {date_to}").font = Font(bold=True, size=13)
    _set_header(ws, 3, ["Дата", "Машин", "Выручка (₽)"])
    for i, row in enumerate(data["daily"], 4):
        ws.cell(row=i, column=1, value=row["date"])
        ws.cell(row=i, column=2, value=row["cars"])
        ws.cell(row=i, column=3, value=round(row["revenue"], 2))

    for c, w in [(1, 14), (2, 10), (3, 16)]:
        ws.column_dimensions[get_column_letter(c)].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"washcontrol_{date_from}_{date_to}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
